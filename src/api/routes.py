# 导入 Flask 相关组件
from flask import Blueprint, request, jsonify, send_file

# 用于生成 Excel 文件
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 导入数据模型
from src.models.student_model import StudentModel
from src.models.course_model import CourseModel
from src.models.score_model import ScoreModel


# 创建蓝图，统一管理 /api 下的所有接口
api_bp = Blueprint("api", __name__, url_prefix="/api")


# =========================
# 通用返回封装
# =========================
def success(data=None, message="success", code=200):
    """
    成功响应封装
    参数：
        data: 返回数据
        message: 提示信息
        code: 状态码
    """
    return jsonify({
        "code": code,
        "message": message,
        "data": data
    }), code


def fail(message="fail", code=400, data=None):
    """
    失败响应封装
    参数：
        message: 错误提示
        code: 状态码
        data: 附加数据
    """
    return jsonify({
        "code": code,
        "message": message,
        "data": data
    }), code


def get_json():
    """
    获取请求体中的 JSON 数据
    如果没有传 JSON，则返回空字典，避免程序报错
    """
    return request.get_json(silent=True) or {}


def create_excel_response(title, headers, rows, filename):
    """
    生成 Excel 文件并返回下载响应
    参数：
        title: 工作表名称
        headers: 表头
        rows: 数据行
        filename: 下载文件名
    """
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # 写入表头
    ws.append(headers)

    # 写入数据内容
    for row in rows:
        ws.append(row)

    # 设置表头样式：蓝色背景、白色字体、居中
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # 自动调整列宽，让导出的 Excel 更美观
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except Exception:
                pass

        ws.column_dimensions[column_letter].width = max_length + 4

    # 将 Excel 文件写入内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 以文件下载形式返回
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# =========================
# 学生 API
# =========================

@api_bp.route("/students", methods=["GET"])
def get_students():
    """获取所有学生信息"""
    students = StudentModel.get_all_students()
    return success(students)


@api_bp.route("/students/<int:student_id>", methods=["GET"])
def get_student_by_id(student_id):
    """根据学生 ID 查询单个学生"""
    student = StudentModel.get_student_by_id(student_id)
    if not student:
        return fail("学生不存在", 404)
    return success(student)


@api_bp.route("/students/search", methods=["GET"])
def search_students():
    """根据关键字搜索学生"""
    keyword = request.args.get("keyword", "").strip()
    if not keyword:
        return fail("keyword 不能为空", 400)

    result = StudentModel.search_students(keyword)
    return success(result)


@api_bp.route("/students/count", methods=["GET"])
def count_students():
    """统计学生总数"""
    total = StudentModel.count_students()
    return success({"total": total})


@api_bp.route("/students", methods=["POST"])
def add_student():
    """新增学生信息"""
    data = get_json()

    # 获取前端传来的学生数据
    student_no = data.get("student_no")
    name = data.get("name")
    gender = data.get("gender", "其他")
    birthday = data.get("birthday")
    phone = data.get("phone")
    email = data.get("email")
    major = data.get("major")
    grade_year = data.get("grade_year")
    status = data.get("status", 1)

    # 学号和姓名是必填项
    if not student_no or not name:
        return fail("student_no 和 name 不能为空", 400)

    try:
        # 调用模型层插入数据
        affected = StudentModel.add_student(
            student_no, name, gender, birthday,
            phone, email, major, grade_year, status
        )

        if affected > 0:
            return success(message="学生添加成功", code=201)
        return fail("学生添加失败", 500)

    except Exception as e:
        return fail(f"学生添加失败：{str(e)}", 500)


@api_bp.route("/students/<int:student_id>", methods=["PUT"])
def update_student(student_id):
    """修改学生信息"""
    data = get_json()

    student_no = data.get("student_no")
    name = data.get("name")
    gender = data.get("gender", "其他")
    birthday = data.get("birthday")
    phone = data.get("phone")
    email = data.get("email")
    major = data.get("major")
    grade_year = data.get("grade_year")
    status = data.get("status", 1)

    if not student_no or not name:
        return fail("student_no 和 name 不能为空", 400)

    try:
        affected = StudentModel.update_student(
            student_id, student_no, name, gender, birthday,
            phone, email, major, grade_year, status
        )

        if affected > 0:
            return success(message="学生修改成功")
        return fail("学生不存在或未更新", 404)

    except Exception as e:
        return fail(f"学生修改失败：{str(e)}", 500)


@api_bp.route("/students/<int:student_id>", methods=["DELETE"])
def delete_student(student_id):
    """删除学生信息"""
    try:
        affected = StudentModel.delete_student(student_id)
        if affected > 0:
            return success(message="学生删除成功")
        return fail("学生不存在或未删除", 404)

    except Exception as e:
        return fail(f"学生删除失败：{str(e)}", 500)


@api_bp.route("/students/<int:student_id>/status", methods=["PATCH"])
def update_student_status(student_id):
    """修改学生状态（启用/禁用）"""
    data = get_json()
    status = data.get("status")

    # 状态只允许 0 或 1
    if status not in [0, 1]:
        return fail("status 只能是 0 或 1", 400)

    try:
        affected = StudentModel.update_status(student_id, status)
        if affected > 0:
            return success(message="学生状态修改成功")
        return fail("学生不存在或状态未修改", 404)

    except Exception as e:
        return fail(f"学生状态修改失败：{str(e)}", 500)


# =========================
# 课程 API
# =========================

@api_bp.route("/courses", methods=["GET"])
def get_courses():
    """获取所有课程信息"""
    courses = CourseModel.get_all_courses()
    return success(courses)


@api_bp.route("/courses/<int:course_id>", methods=["GET"])
def get_course_by_id(course_id):
    """根据课程 ID 查询单条课程信息"""
    course = CourseModel.get_course_by_id(course_id)
    if not course:
        return fail("课程不存在", 404)
    return success(course)


@api_bp.route("/courses/search", methods=["GET"])
def search_courses():
    """根据关键字搜索课程"""
    keyword = request.args.get("keyword", "").strip()
    if not keyword:
        return fail("keyword 不能为空", 400)

    result = CourseModel.search_courses(keyword)
    return success(result)


@api_bp.route("/courses/count", methods=["GET"])
def count_courses():
    """统计课程总数"""
    total = CourseModel.count_courses()
    return success({"total": total})


@api_bp.route("/courses", methods=["POST"])
def add_course():
    """新增课程"""
    data = get_json()

    course_no = data.get("course_no")
    course_name = data.get("course_name")
    credit = data.get("credit", 0.0)
    teacher = data.get("teacher")
    description = data.get("description")
    status = data.get("status", 1)

    if not course_no or not course_name:
        return fail("course_no 和 course_name 不能为空", 400)

    try:
        affected = CourseModel.add_course(
            course_no, course_name, credit, teacher, description, status
        )

        if affected > 0:
            return success(message="课程添加成功", code=201)
        return fail("课程添加失败", 500)

    except Exception as e:
        return fail(f"课程添加失败：{str(e)}", 500)


@api_bp.route("/courses/<int:course_id>", methods=["PUT"])
def update_course(course_id):
    """修改课程信息"""
    data = get_json()

    course_no = data.get("course_no")
    course_name = data.get("course_name")
    credit = data.get("credit", 0.0)
    teacher = data.get("teacher")
    description = data.get("description")
    status = data.get("status", 1)

    if not course_no or not course_name:
        return fail("course_no 和 course_name 不能为空", 400)

    try:
        affected = CourseModel.update_course(
            course_id, course_no, course_name, credit, teacher, description, status
        )

        if affected > 0:
            return success(message="课程修改成功")
        return fail("课程不存在或未更新", 404)

    except Exception as e:
        return fail(f"课程修改失败：{str(e)}", 500)


@api_bp.route("/courses/<int:course_id>", methods=["DELETE"])
def delete_course(course_id):
    """删除课程"""
    try:
        affected = CourseModel.delete_course(course_id)
        if affected > 0:
            return success(message="课程删除成功")
        return fail("课程不存在或未删除", 404)

    except Exception as e:
        return fail(f"课程删除失败：{str(e)}", 500)


@api_bp.route("/courses/<int:course_id>/status", methods=["PATCH"])
def update_course_status(course_id):
    """修改课程状态（启用/禁用）"""
    data = get_json()
    status = data.get("status")

    if status not in [0, 1]:
        return fail("status 只能是 0 或 1", 400)

    try:
        affected = CourseModel.update_status(course_id, status)
        if affected > 0:
            return success(message="课程状态修改成功")
        return fail("课程不存在或状态未修改", 404)

    except Exception as e:
        return fail(f"课程状态修改失败：{str(e)}", 500)


# =========================
# 成绩 API
# =========================

@api_bp.route("/scores", methods=["GET"])
def get_scores():
    """获取所有成绩信息"""
    scores = ScoreModel.get_all_scores()
    return success(scores)


@api_bp.route("/scores/<int:score_id>", methods=["GET"])
def get_score_by_id(score_id):
    """根据成绩 ID 查询成绩"""
    score = ScoreModel.get_score_by_id(score_id)
    if not score:
        return fail("成绩不存在", 404)
    return success(score)


@api_bp.route("/scores/search", methods=["GET"])
def search_scores():
    """根据关键字搜索成绩"""
    keyword = request.args.get("keyword", "").strip()
    if not keyword:
        return fail("keyword 不能为空", 400)

    result = ScoreModel.search_scores(keyword)
    return success(result)


@api_bp.route("/scores/count", methods=["GET"])
def count_scores():
    """统计成绩总数"""
    total = ScoreModel.count_scores()
    return success({"total": total})


@api_bp.route("/scores/rankings", methods=["GET"])
def get_rankings():
    """获取成绩排名数据，可按课程、学期、数量限制筛选"""
    course_id = request.args.get("course_id", type=int)
    term = request.args.get("term")
    limit = request.args.get("limit", type=int)

    result = ScoreModel.get_rankings(course_id=course_id, term=term, limit=limit)
    return success(result)


@api_bp.route("/scores/statistics/student/<int:student_id>", methods=["GET"])
def statistics_by_student(student_id):
    """查询某个学生的成绩统计"""
    result = ScoreModel.get_statistics_by_student(student_id)
    if not result:
        return fail("没有找到该学生的成绩统计", 404)
    return success(result)


@api_bp.route("/scores/statistics/course/<int:course_id>", methods=["GET"])
def statistics_by_course(course_id):
    """查询某门课程的成绩统计"""
    result = ScoreModel.get_statistics_by_course(course_id)
    if not result:
        return fail("没有找到该课程的成绩统计", 404)
    return success(result)


@api_bp.route("/scores", methods=["POST"])
def add_score():
    """新增成绩"""
    data = get_json()

    student_id = data.get("student_id")
    course_id = data.get("course_id")
    usual_score = data.get("usual_score", 0)
    exam_score = data.get("exam_score", 0)
    total_score = data.get("total_score")
    term = data.get("term")
    remark = data.get("remark")

    if not student_id or not course_id:
        return fail("student_id 和 course_id 不能为空", 400)

    try:
        affected = ScoreModel.add_score(
            student_id, course_id, usual_score, exam_score,
            total_score, term, remark
        )

        if affected > 0:
            return success(message="成绩添加成功", code=201)
        return fail("成绩添加失败", 500)

    except Exception as e:
        return fail(f"成绩添加失败：{str(e)}", 500)


@api_bp.route("/scores/<int:score_id>", methods=["PUT"])
def update_score(score_id):
    """修改成绩"""
    data = get_json()

    student_id = data.get("student_id")
    course_id = data.get("course_id")
    usual_score = data.get("usual_score", 0)
    exam_score = data.get("exam_score", 0)
    total_score = data.get("total_score")
    term = data.get("term")
    remark = data.get("remark")

    if not student_id or not course_id:
        return fail("student_id 和 course_id 不能为空", 400)

    try:
        affected = ScoreModel.update_score(
            score_id, student_id, course_id,
            usual_score, exam_score, total_score, term, remark
        )

        if affected > 0:
            return success(message="成绩修改成功")
        return fail("成绩不存在或未更新", 404)

    except Exception as e:
        return fail(f"成绩修改失败：{str(e)}", 500)


@api_bp.route("/scores/<int:score_id>", methods=["DELETE"])
def delete_score(score_id):
    """删除成绩"""
    try:
        affected = ScoreModel.delete_score(score_id)
        if affected > 0:
            return success(message="成绩删除成功")
        return fail("成绩不存在或未删除", 404)

    except Exception as e:
        return fail(f"成绩删除失败：{str(e)}", 500)


@api_bp.route("/scores/student/<int:student_id>", methods=["GET"])
def get_scores_by_student(student_id):
    """查询某个学生的所有成绩"""
    result = ScoreModel.get_scores_by_student(student_id)
    return success(result)


@api_bp.route("/scores/course/<int:course_id>", methods=["GET"])
def get_scores_by_course(course_id):
    """查询某门课程的所有成绩"""
    result = ScoreModel.get_scores_by_course(course_id)
    return success(result)


# =========================
# 导出 Excel 接口
# =========================

@api_bp.route("/export/students", methods=["GET"])
def export_students():
    """导出学生信息到 Excel"""
    students = StudentModel.get_all_students()

    # Excel 表头
    headers = ["ID", "学号", "姓名", "性别", "出生日期", "手机号", "邮箱", "专业", "年级", "状态"]

    # 整理导出数据
    rows = []
    for s in students:
        rows.append([
            s.get("id"),
            s.get("student_no"),
            s.get("name"),
            s.get("gender"),
            s.get("birthday"),
            s.get("phone"),
            s.get("email"),
            s.get("major"),
            s.get("grade_year"),
            "正常" if s.get("status") == 1 else "禁用"
        ])

    return create_excel_response(
        title="学生信息",
        headers=headers,
        rows=rows,
        filename="学生信息导出.xlsx"
    )


@api_bp.route("/export/scores", methods=["GET"])
def export_scores():
    """导出成绩信息到 Excel"""
    scores = ScoreModel.get_all_scores()

    # Excel 表头
    headers = ["ID", "学号", "姓名", "课程编号", "课程名称", "平时成绩", "期末成绩", "总评成绩", "学期", "备注"]

    # 整理导出数据
    rows = []
    for s in scores:
        rows.append([
            s.get("id"),
            s.get("student_no"),
            s.get("student_name"),
            s.get("course_no"),
            s.get("course_name"),
            s.get("usual_score"),
            s.get("exam_score"),
            s.get("total_score"),
            s.get("term"),
            s.get("remark")
        ])

    return create_excel_response(
        title="成绩信息",
        headers=headers,
        rows=rows,
        filename="成绩信息导出.xlsx"
    )
